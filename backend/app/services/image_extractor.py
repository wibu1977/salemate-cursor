"""Phase 2: Extract embedded images from XLSX files and map them to rows.

XLSX files are ZIP archives with images stored under ``xl/media/``.
We use ``openpyxl`` (already a dependency) to load the workbook in
data-only mode and inspect each worksheet's ``_images`` list to resolve
the anchor row for every image.  If the structured anchor lookup fails
(some Excel builds use different anchor types), we fall back to
sequential order (image 1 → data row 1, image 2 → data row 2, …).

The companion ``upload_and_map`` helper pushes each extracted image
to Cloudinary via ``StorageService`` and returns
``{data_row_index: cloudinary_url}``.

*worksheet_title*: when set, read images only from that tab (recommended
when the XLSX has multiple worksheets, e.g. Google Drive spreadsheet export).
"""

from __future__ import annotations

import io
import logging
import zipfile
from typing import Any

logger = logging.getLogger("salemate.image_extractor")


def _select_worksheet(wb: Any, worksheet_title: str | None) -> Any:
    """Pick worksheet tab (exported Google Sheets tab names align with workbook sheet titles)."""
    if worksheet_title and str(worksheet_title).strip():
        t = str(worksheet_title).strip()
        if t in wb.sheetnames:
            return wb[t]
        tl = t.lower()
        for name in wb.sheetnames:
            if str(name).strip().lower() == tl:
                return wb[name]
    return wb.active or wb.worksheets[0]


def _image_bytes(img: Any, raw_media: dict[str, bytes]) -> bytes | None:
    """Resolve drawing image object to ``xl/media/`` bytes (or PIL fallback)."""
    ref = getattr(img, "path", None) or getattr(img, "ref", "")
    candidates = [ref, f"xl/{ref}", ref.replace("../", "xl/")]
    img_bytes: bytes | None = None
    for c in candidates:
        if c in raw_media:
            img_bytes = raw_media[c]
            break

    if img_bytes is None:
        try:
            buf = io.BytesIO()
            img.image.save(buf, format=img.image.format or "PNG")
            img_bytes = buf.getvalue()
        except Exception:
            return None
    return img_bytes


# ---------------------------------------------------------------------------
# Low-level extraction
# ---------------------------------------------------------------------------

def extract_images_from_xlsx(
    content: bytes,
    *,
    data_start_row_0based: int = 1,
    worksheet_title: str | None = None,
) -> dict[int, bytes]:
    """Return ``{data_row_0based: image_bytes}`` for every embedded image.

    Steps:
        1.  Open the XLSX as a ``zipfile`` and collect raw bytes for every
            file in ``xl/media/``.
        2.  Use ``openpyxl.load_workbook`` (non-read_only, so that the
            ``_images`` attribute is populated) to walk each worksheet's
            images and resolve their anchor row.
        3.  If structured anchors are missing, fall back to sequential
            mapping starting at *data_start_row_0based*.
    """
    # Quick sanity: must be a ZIP / valid XLSX
    bio = io.BytesIO(content)
    if not zipfile.is_zipfile(bio):
        return {}
    bio.seek(0)

    # ----- Gather raw bytes from xl/media/ ----------------------------
    raw_media: dict[str, bytes] = {}
    try:
        with zipfile.ZipFile(bio, "r") as zf:
            for name in zf.namelist():
                if name.startswith("xl/media/"):
                    raw_media[name] = zf.read(name)
    except Exception:
        logger.warning("Failed to read xl/media/ from XLSX ZIP", exc_info=True)
        return {}

    if not raw_media:
        return {}

    # ----- Try openpyxl anchor-based mapping --------------------------
    bio.seek(0)
    try:
        import openpyxl  # type: ignore[import-untyped]

        wb = openpyxl.load_workbook(bio, data_only=True)
        try:
            ws = _select_worksheet(wb, worksheet_title)

            mapped: dict[int, bytes] = {}
            for img in getattr(ws, "_images", []):
                anchor = getattr(img, "anchor", None)
                if anchor is None:
                    continue
                from_marker = getattr(anchor, "_from", None) or getattr(anchor, "anchorFrom", None)
                if from_marker is None:
                    continue
                row_0 = getattr(from_marker, "row", None)
                if row_0 is None:
                    continue
                row_0 = int(row_0)

                img_bytes = _image_bytes(img, raw_media)
                if img_bytes and row_0 not in mapped:
                    mapped[row_0] = img_bytes

            if mapped:
                logger.info(
                    "Anchor-based extraction succeeded: %d images mapped",
                    len(mapped),
                )
                return mapped

            ws_images = list(getattr(ws, "_images", []) or [])
            seq: dict[int, bytes] = {}
            for idx, img in enumerate(ws_images):
                b = _image_bytes(img, raw_media)
                if b:
                    seq[data_start_row_0based + idx] = b

            if seq:
                logger.info(
                    "Sheet sequential image extraction: %d images → rows from %d",
                    len(seq),
                    data_start_row_0based,
                )
                return seq
        finally:
            wb.close()

    except Exception:
        logger.info(
            "Anchor-based image mapping failed, falling back to sequential",
            exc_info=True,
        )

    # ----- Fallback: sequential mapping --------------------------------
    # Sort media files by name (image1.png, image2.png, …) and assign to
    # consecutive data rows.
    sorted_names = sorted(raw_media.keys())
    sequential: dict[int, bytes] = {}
    for idx, name in enumerate(sorted_names):
        row_0 = data_start_row_0based + idx
        sequential[row_0] = raw_media[name]

    logger.info(
        "Sequential fallback: %d images → data rows starting at %d",
        len(sequential),
        data_start_row_0based,
    )
    return sequential


# ---------------------------------------------------------------------------
# Upload helper
# ---------------------------------------------------------------------------

async def upload_and_map(
    images: dict[int, bytes],
    workspace_id: str,
) -> dict[int, str]:
    """Upload each image to Cloudinary and return ``{row_0based: url}``."""
    from app.services.storage_service import StorageService  # deferred to avoid circular

    url_map: dict[int, str] = {}
    for row_idx, img_bytes in images.items():
        # Skip very large images (> 10 MB raw)
        if len(img_bytes) > 10 * 1024 * 1024:
            logger.warning("Skipping oversized image for row %d (%d bytes)", row_idx, len(img_bytes))
            continue
        try:
            url = await StorageService.upload_product_image(
                img_bytes, workspace_id, row_idx
            )
            if url:
                url_map[row_idx] = url
        except Exception:
            logger.warning("Failed to upload image for row %d", row_idx, exc_info=True)

    logger.info("Uploaded %d/%d images to Cloudinary", len(url_map), len(images))
    return url_map
