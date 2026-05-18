"""Đọc Google Sheets (OAuth) / file CSV-Excel và import sản phẩm / khách hàng — validation + upsert + dry-run."""

from __future__ import annotations

import asyncio
import csv
import io
import logging
import re
import uuid
from difflib import SequenceMatcher
from typing import Any, Literal

import httpx
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.customer import Customer
from app.models.inventory import Product
from app.models.workspace import Workspace
from app.services.google_oauth_service import ensure_valid_access_token
from app.services.notification_service import NotificationService

logger = logging.getLogger("salemate.sheets_import")

DuplicateStrategy = Literal["skip", "update", "create_all"]

# Khớp tiêu đề cột (không phân biệt hoa thường, có dấu / không dấu đơn giản)
NAME_KEYS = (
    "name",
    "tên",
    "ten",
    "product",
    "sản phẩm",
    "san pham",
    "tên sản phẩm",
)
FIRST_NAME_KEYS = (
    "first_name",
    "first name",
    "given name",
    "tên riêng",
    "ten rieng",
    "tên",
    "ten",
)
LAST_NAME_KEYS = (
    "last_name",
    "last name",
    "surname",
    "family name",
    "họ",
    "ho",
    "đệm",
    "dem",
)
PRICE_KEYS = ("price", "giá", "gia", "gia ban")
QTY_KEYS = ("quantity", "số lượng", "so luong", "stock", "ton", "tồn")
THRESHOLD_KEYS = ("threshold", "ngưỡng", "nguong", "stock_threshold")
DESC_KEYS = ("description", "mô tả", "mo ta")
IMAGE_KEYS = ("image", "ảnh", "anh", "image_url", "hình", "hinh")
CATEGORY_KEYS = ("category", "danh mục", "danh muc", "type", "loại", "loai", "nhóm", "nhom")
PHONE_KEYS = ("phone", "sđt", "sdt", "tel", "mobile", "điện thoại", "dien thoai")
EMAIL_KEYS = ("email", "e-mail", "mail")
NOTE_KEYS = ("note", "ghi chú", "ghichu", "ghi chu", "comments")

PRODUCT_FIELD_SPECS: dict[str, tuple[str, ...]] = {
    "name": NAME_KEYS,
    "price": PRICE_KEYS,
    "quantity": QTY_KEYS,
    "stock_threshold": THRESHOLD_KEYS,
    "description": DESC_KEYS,
    "image_url": IMAGE_KEYS,
    "category": CATEGORY_KEYS,
    "first_name": FIRST_NAME_KEYS,
    "last_name": LAST_NAME_KEYS,
}

# Smart defaults for missing columns — Phase 1: auto-fill
FIELD_DEFAULTS: dict[str, int | str] = {
    "quantity": 0,           # "Chưa nhập số lượng"
    "stock_threshold": 5,    # Global safe default
    "category": "",          # Will be filled by AI categorizer (Phase 3)
    "description": "",       # Optional free-text
    "image_url": "",         # Will be filled by image extractor (Phase 2)
}

CUSTOMER_FIELD_SPECS: dict[str, tuple[str, ...]] = {
    "name": NAME_KEYS + ("họ tên", "ho ten", "fullname", "full name"),
    "phone": PHONE_KEYS,
    "email": EMAIL_KEYS,
    "note": NOTE_KEYS,
    "first_name": FIRST_NAME_KEYS,
    "last_name": LAST_NAME_KEYS,
}


def _norm_key(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _build_header_index(headers: list[str]) -> dict[str, int]:
    idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        k = _norm_key(str(h))
        if k:
            idx[k] = i
    return idx


def _fuzzy_ratio(a: str, b: str) -> float:
    return SequenceMatcher(None, _norm_key(a), _norm_key(b)).ratio()


def suggest_field_for_headers(
    field_key: str,
    headers: list[str],
    key_map: tuple[str, ...],
) -> tuple[str | None, float]:
    """Gợi ý cột tốt nhất cho field_key; trả về (header gốc hoặc None, confidence 0–1)."""
    best_h: str | None = None
    best_score = 0.0
    for h in headers:
        hs = str(h or "").strip()
        if not hs:
            continue
        for key in key_map:
            nk = _norm_key(key)
            hn = _norm_key(hs)
            if nk == hn:
                return hs, 1.0
            if nk and (nk in hn or hn in nk):
                sc = 0.92
                if sc > best_score:
                    best_score = sc
                    best_h = hs
                break
            r = _fuzzy_ratio(nk, hn) if nk else 0.0
            if r > 0.72 and r > best_score:
                best_score = r
                best_h = hs
    return best_h, best_score


def suggest_column_mapping(
    headers: list[str],
    entity: str,
) -> dict[str, dict[str, Any]]:
    """Gợi ý mapping { field_key: { \"header\": str|None, \"confidence\": float } }."""
    entity_norm = (entity or "products").lower()
    specs = CUSTOMER_FIELD_SPECS if entity_norm in ("customer", "customers", "khach", "khách") else PRODUCT_FIELD_SPECS
    out: dict[str, dict[str, Any]] = {}
    used: set[str] = set()
    for fk, keys in specs.items():
        hdr, conf = suggest_field_for_headers(fk, headers, keys)
        if hdr and hdr in used:
            # tránh trùng cột cho 2 field
            out[fk] = {"header": None, "confidence": 0.0}
            continue
        if hdr:
            used.add(hdr)
        out[fk] = {"header": hdr, "confidence": round(conf, 3)}
    return out


def _find_col(
    header_idx: dict[str, int],
    key_map: tuple[str, ...],
    mapping_key: str | None = None,
    custom_mapping: dict[str, str] | None = None,
) -> int | None:
    if custom_mapping and mapping_key and mapping_key in custom_mapping:
        target_header = _norm_key(custom_mapping[mapping_key])
        if target_header in header_idx:
            return header_idx[target_header]

    for key in key_map:
        nk = _norm_key(key)
        if nk in header_idx:
            return header_idx[nk]
    for hk, col in header_idx.items():
        for key in key_map:
            nk = _norm_key(key)
            if nk and (nk in hk or hk in nk):
                return col
    return None


def _cell(row: list[Any], i: int | None) -> str:
    if i is None or i >= len(row):
        return ""
    v = row[i]
    if v is None:
        return ""
    return str(v).strip()


def _mapped_header_names(column_mapping: dict[str, str] | None) -> set[str]:
    if not column_mapping:
        return set()
    return {str(v).strip() for v in column_mapping.values() if str(v).strip()}


def _format_extra_cell(raw: str) -> str:
    t = (raw or "").strip()
    if not t:
        return t
    try:
        from dateutil import parser as date_parser

        dt = date_parser.parse(t, dayfirst=False, fuzzy=True)
        return dt.date().isoformat()
    except (ValueError, TypeError, OverflowError):
        return t


def normalize_product_name(raw: str) -> str:
    return re.sub(r"\s+", " ", (raw or "").strip())


def _row_import_extra(
    row: list[Any],
    headers: list[str],
    column_mapping: dict[str, str] | None,
) -> dict[str, str]:
    """Cột chưa ánh xạ → lưu dưới import_extra (ghi đè theo header gốc)."""
    mapped = _mapped_header_names(column_mapping)
    extra: dict[str, str] = {}
    for i, h in enumerate(headers):
        label = str(h or "").strip()
        if not label or label in mapped:
            continue
        val = _cell(row, i)
        if val:
            extra[label] = _format_extra_cell(val)
    return extra


_SAFE_EMAIL = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def normalize_email(raw: str) -> str | None:
    s = (raw or "").strip()
    if not s:
        return None
    if _SAFE_EMAIL.match(s):
        return s.lower()
    return None


def parse_price(raw: str | int | float) -> int:
    """Bỏ ký hiệu tiền, xử lý phân cách hàng nghìn / thập phân (.,)."""
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return int(round(float(raw)))
    s = str(raw or "").strip()
    if not s:
        return 0
    s = re.sub(r"^[\s₩$€£¥]+\s*", "", s)
    s = re.sub(r"\s*KRW\b|\s*VND\b|\s*USD\b|\s*원\b", "", s, flags=re.I)
    s = s.replace("\u00a0", "").replace(" ", "")
    neg = s.startswith("-")
    if neg:
        s = s[1:]
    core = re.sub(r"[^\d.,]", "", s)
    if not core:
        return 0
    last_comma = core.rfind(",")
    last_dot = core.rfind(".")
    normalized: str
    if last_comma != -1 and last_dot != -1:
        if last_comma > last_dot:
            normalized = core.replace(".", "").replace(",", ".")
        else:
            normalized = core.replace(",", "")
    elif last_comma != -1 and last_dot == -1:
        parts = core.split(",")
        if len(parts[-1]) in (3,) and all(len(p) <= 3 for p in parts):
            normalized = "".join(parts)
        elif len(parts[-1]) == 2 and parts[0]:
            normalized = core.replace(",", ".")
        else:
            normalized = core.replace(",", "")
    else:
        normalized = core.replace(",", "") if last_dot != -1 else core.replace(",", ".")
    try:
        v = float(normalized or 0)
    except ValueError:
        v = 0.0
    iv = int(round(abs(v)))
    return -iv if neg else iv


def normalize_phone(raw: str) -> str | None:
    """Chuẩn hóa E.164 tối thiểu: VN (+84), KR (+82), còn lại +digits nếu đủ dài."""
    s = (raw or "").strip()
    if not s:
        return None
    if s.startswith("+"):
        d = re.sub(r"\D", "", s[1:])
        if len(d) < 8:
            return None
        return "+" + d
    digits = re.sub(r"\D", "", s)
    if not digits:
        return None
    if s.startswith("00"):
        digits = digits[2:] if len(digits) > 2 else digits
        if len(digits) >= 8:
            return "+" + digits
    if len(digits) >= 10 and digits.startswith("010"):
        return "+82" + digits[1:]
    if len(digits) == 10 and digits.startswith("10"):
        return "+82" + digits
    if len(digits) >= 11 and digits.startswith("82"):
        return "+" + digits
    if len(digits) >= 10 and digits.startswith("84"):
        return "+" + digits
    if digits.startswith("0") and not digits.startswith("00"):
        return "+84" + digits[1:]
    if len(digits) >= 8:
        return "+" + digits
    return None


def _concat_customer_name(
    row: list[Any],
    col_name: int | None,
    col_first: int | None,
    col_last: int | None,
    column_mapping: dict[str, str] | None,
) -> str:
    if col_name is not None:
        n = _cell(row, col_name)
        if n:
            return n
    parts: list[str] = []
    if col_first is not None:
        parts.append(_cell(row, col_first))
    if col_last is not None:
        parts.append(_cell(row, col_last))
    merged = " ".join(p for p in parts if p).strip()
    return merged


async def _get_sheets_service(workspace: Workspace, access_token: str):
    def _build():
        creds = Credentials(token=access_token)
        return build("sheets", "v4", credentials=creds, cache_discovery=False)

    return await asyncio.to_thread(_build)


def _extract_spreadsheet_id(raw: str) -> str:
    raw = (raw or "").strip()
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", raw)
    if m:
        return m.group(1)
    return raw


def guess_header_data_rows_1based(rows: list[list[Any]], entity: str) -> tuple[int, int]:
    """Gợi ý dòng tiêu đề và dòng bắt đầu dữ liệu (1-based), mặc định 1 và 2."""
    ent = (entity or "products").lower()
    is_customer = ent in ("customer", "customers", "khach", "khách")
    for i in range(min(10, len(rows))):
        headers = [str(h or "") for h in rows[i]]
        hidx = _build_header_index(headers)
        if is_customer:
            if _find_col(hidx, PHONE_KEYS, None, None) is not None or _find_col(
                hidx, EMAIL_KEYS, None, None
            ) is not None:
                return i + 1, i + 2
        else:
            if _find_col(hidx, NAME_KEYS, None, None) is not None:
                return i + 1, i + 2
    return 1, 2


def analyze_file_structure(
    content: bytes,
    filename: str,
    entity: str = "products",
) -> dict[str, Any]:
    """Analyze a file's structure: detect columns, missing fields, smart defaults, embedded images.

    Returns a dict suitable for the ImportAnalysisResponse schema.
    """
    rows = parse_upload_to_rows(content, filename, max_rows=200)
    if not rows:
        return {
            "matched_fields": {},
            "missing_fields": [],
            "defaults_applied": {},
            "total_rows": 0,
            "has_embedded_images": False,
            "embedded_image_count": 0,
        }

    # detect header row
    h1, d1 = guess_header_data_rows_1based(rows, entity)
    header_idx = max(0, h1 - 1)
    headers = [str(h or "") for h in rows[header_idx]] if header_idx < len(rows) else []

    # fuzzy column matching
    mapping = suggest_column_mapping(headers, entity)
    matched: dict[str, dict[str, Any]] = {}
    missing: list[str] = []

    # required + optional fields for products
    product_fields = {"name", "price", "quantity", "stock_threshold", "description", "image_url", "category"}

    for field in product_fields:
        entry = mapping.get(field, {"header": None, "confidence": 0.0})
        if entry.get("header") and entry.get("confidence", 0) > 0.5:
            matched[field] = {"header": entry["header"], "confidence": entry["confidence"]}
        else:
            missing.append(field)

    # smart defaults for missing fields (only optional ones)
    defaults_applied: dict[str, int | str] = {}
    for f in missing:
        if f in FIELD_DEFAULTS:
            defaults_applied[f] = FIELD_DEFAULTS[f]

    # detect embedded images in XLSX
    has_images = False
    image_count = 0
    name_lower = (filename or "").lower()
    if name_lower.endswith(".xlsx"):
        try:
            import zipfile as _zf
            bio = io.BytesIO(content)
            with _zf.ZipFile(bio, "r") as zf:
                media_files = [n for n in zf.namelist() if n.startswith("xl/media/")]
                image_count = len(media_files)
                has_images = image_count > 0
        except Exception:
            pass  # not a valid zip / corrupted — skip silently

    # count data rows (excluding header)
    all_rows = parse_upload_to_rows(content, filename, max_rows=None)
    data_rows = max(0, len(all_rows) - (header_idx + 1))

    return {
        "matched_fields": matched,
        "missing_fields": missing,
        "defaults_applied": defaults_applied,
        "total_rows": data_rows,
        "has_embedded_images": has_images,
        "embedded_image_count": image_count,
    }


def parse_upload_to_rows(content: bytes, filename: str, max_rows: int | None = None) -> list[list[Any]]:
    """Parse .csv, .xlsx, .xls thành rows_2d giống Google Sheets API."""
    name = (filename or "").lower()
    rows: list[list[Any]] = []
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        for i, r in enumerate(reader):
            rows.append([c if c != "" else "" for c in r])
            if max_rows is not None and len(rows) >= max_rows:
                break
        return rows
    if name.endswith(".xlsx"):
        import openpyxl

        bio = io.BytesIO(content)
        wb = openpyxl.load_workbook(bio, read_only=True, data_only=True)
        try:
            ws = wb.active
            limit = max_rows
            for r in ws.iter_rows(values_only=True):
                rows.append([("" if c is None else c) for c in r])
                if limit is not None and len(rows) >= limit:
                    break
        finally:
            wb.close()
        return rows
    if name.endswith(".xls"):
        import xlrd

        book = xlrd.open_workbook(file_contents=content)
        sheet = book.sheet_by_index(0)
        n = sheet.nrows if max_rows is None else min(sheet.nrows, max_rows)
        for ri in range(n):
            row: list[Any] = []
            for ci in range(sheet.ncols):
                cell = sheet.cell(ri, ci)
                v = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE and v:
                    try:
                        t = xlrd.xldate_as_datetime(v, book.datemode)
                        row.append(t.strftime("%Y-%m-%d"))
                    except Exception:
                        row.append(str(v))
                else:
                    row.append("" if v is None else v)
            rows.append(row)
        return rows
    raise ValueError("Định dạng không hỗ trợ. Dùng .csv, .xlsx hoặc .xls.")


async def list_sheet_titles(db: AsyncSession, workspace: Workspace, spreadsheet_id: str) -> list[str]:
    spreadsheet_id = _extract_spreadsheet_id(spreadsheet_id)
    access = await ensure_valid_access_token(db, workspace)
    service = await _get_sheets_service(workspace, access)

    def _call():
        meta = (
            service.spreadsheets()
            .get(spreadsheetId=spreadsheet_id, fields="sheets(properties(title))")
            .execute()
        )
        sheets = meta.get("sheets") or []
        return [s.get("properties", {}).get("title") or "" for s in sheets]

    try:
        return await asyncio.to_thread(_call)
    except HttpError as e:
        logger.warning("Sheets metadata error: %s", e)
        raise ValueError(_http_error_message(e)) from e


async def fetch_sheet_preview(
    db: AsyncSession,
    workspace: Workspace,
    spreadsheet_id: str,
    sheet_name: str,
    limit: int = 10,
) -> list[list[Any]]:
    spreadsheet_id = _extract_spreadsheet_id(spreadsheet_id)
    access = await ensure_valid_access_token(db, workspace)
    service = await _get_sheets_service(workspace, access)

    rng = f"'{sheet_name.replace(chr(39), chr(39) * 2)}'!1:{limit}"

    def _call():
        resp = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=spreadsheet_id, range=rng, majorDimension="ROWS")
            .execute()
        )
        return resp.get("values") or []

    try:
        return await asyncio.to_thread(_call)
    except HttpError as e:
        raise ValueError(_http_error_message(e)) from e


async def fetch_sheet_values(
    db: AsyncSession,
    workspace: Workspace,
    spreadsheet_id: str,
    sheet_name: str,
    range_a1: str | None = None,
) -> list[list[Any]]:
    spreadsheet_id = _extract_spreadsheet_id(spreadsheet_id)
    access = await ensure_valid_access_token(db, workspace)
    service = await _get_sheets_service(workspace, access)
    rng = f"'{sheet_name.replace(chr(39), chr(39) * 2)}'!{range_a1}" if range_a1 else f"'{sheet_name.replace(chr(39), chr(39) * 2)}'"

    def _call():
        resp = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=spreadsheet_id, range=rng, majorDimension="ROWS")
            .execute()
        )
        return resp.get("values") or []

    try:
        return await asyncio.to_thread(_call)
    except HttpError as e:
        raise ValueError(_http_error_message(e)) from e


_XLSX_EXPORT_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


async def fetch_spreadsheet_export_xlsx_bytes(
    db: AsyncSession,
    workspace: Workspace,
    spreadsheet_id: str,
    *,
    timeout_s: float = 120.0,
) -> bytes:
    """Export a Google Spreadsheet as XLSX via Drive API (embedded images → ``xl/media``)."""
    sid = _extract_spreadsheet_id(spreadsheet_id)
    access = await ensure_valid_access_token(db, workspace)
    export_url = f"https://www.googleapis.com/drive/v3/files/{sid}/export"

    async with httpx.AsyncClient() as client:
        r = await client.get(
            export_url,
            params={"mimeType": _XLSX_EXPORT_MIME},
            headers={"Authorization": f"Bearer {access}"},
            timeout=timeout_s,
        )
        if r.status_code == 403:
            logger.warning(
                "Drive export forbidden spreadsheet_id=%s body=%s", sid, r.text[:400]
            )
            raise ValueError(
                "Google từ chối export bảng tính để đọc ảnh. Hãy ngắt kết nối và "
                "kết nối lại Google (cần quyền đọc Drive), và bật Google Drive API trên GCP."
            ) from None
        if r.status_code >= 400:
            logger.warning(
                "Drive export failed %s spreadsheet_id=%s body=%s",
                r.status_code,
                sid,
                r.text[:400],
            )
            raise ValueError(
                f"Export Google Sheet sang XLSX thất bại (HTTP {r.status_code})."
            ) from None
        content = r.content

    if len(content) < 64:
        raise ValueError("Export Google Sheet trả dữ liệu quá nhỏ — không hợp lệ.")
    return content


def _http_error_message(e: HttpError) -> str:
    msg = str(e)
    if e.resp.status == 403:
        return "Google từ chối truy cập. Đảm bảo đã chọn đúng file và tài khoản có quyền xem."
    if e.resp.status == 404:
        return "Không tìm thấy spreadsheet hoặc sheet."
    return f"Lỗi Google Sheets API: {msg}"


def build_import_error_csv(records: list[dict[str, Any]]) -> str:
    if not records:
        return ""
    fieldnames = list(records[0].keys())
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    w.writeheader()
    for r in records:
        w.writerow({k: r.get(k, "") for k in fieldnames})
    return buf.getvalue()


async def _after_import_commit(db: AsyncSession, dry_run: bool, workspace_id: uuid.UUID) -> None:
    if dry_run:
        await db.rollback()
        return
    await db.commit()
    try:
        from app.workers.embeddings import embed_workspace_products

        embed_workspace_products.delay(str(workspace_id))
    except Exception as e:
        logger.warning("embed_workspace_products: %s", e)


async def import_products_from_values(
    db: AsyncSession,
    workspace_id: uuid.UUID,
    rows_2d: list[list[Any]],
    header_row_index: int = 0,
    data_start_index: int = 1,
    column_mapping: dict[str, str] | None = None,
    *,
    duplicate_strategy: DuplicateStrategy = "update",
    dry_run: bool = False,
    error_sample_limit: int = 50,
    extracted_images: dict[int, str] | None = None,
    ai_categories: dict[str, str] | None = None,
    default_overrides: dict[str, Any] | None = None,
    manual_overrides: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    if not rows_2d:
        raise ValueError("Sheet trống hoặc không có dữ liệu.")

    current_h_idx = header_row_index
    best_hidx: dict[str, int] = {}
    col_name: int | None = None

    for i in range(current_h_idx, min(len(rows_2d), current_h_idx + 10)):
        headers = [str(h or "") for h in rows_2d[i]]
        hidx = _build_header_index(headers)
        test_name = _find_col(hidx, NAME_KEYS, "name", column_mapping)
        if test_name is not None:
            col_name = test_name
            best_hidx = hidx
            header_row_index = i
            if data_start_index <= header_row_index:
                data_start_index = i + 1
            break

    if col_name is None:
        raise ValueError(
            "Không tìm thấy cột tên sản phẩm (hệ thống không tự đoán được dòng tiêu đề, vui lòng kiểm tra lại file)."
        )

    hidx = best_hidx
    headers_row = [str(h or "") for h in rows_2d[header_row_index]]
    col_price = _find_col(hidx, PRICE_KEYS, "price", column_mapping)
    col_qty = _find_col(hidx, QTY_KEYS, "quantity", column_mapping)
    col_thr = _find_col(hidx, THRESHOLD_KEYS, "stock_threshold", column_mapping)
    col_desc = _find_col(hidx, DESC_KEYS, "description", column_mapping)
    col_image = _find_col(hidx, IMAGE_KEYS, "image_url", column_mapping)
    col_cat = _find_col(hidx, CATEGORY_KEYS, "category", column_mapping)
    col_first = _find_col(hidx, FIRST_NAME_KEYS, "first_name", column_mapping)
    col_last = _find_col(hidx, LAST_NAME_KEYS, "last_name", column_mapping)

    # --- Phase 1: detect missing columns & track auto-fills ---
    _col_lookup = {
        "price": col_price, "quantity": col_qty, "stock_threshold": col_thr,
        "description": col_desc, "image_url": col_image, "category": col_cat,
    }
    missing_fields: list[str] = [f for f, c in _col_lookup.items() if c is None]
    
    # Merge global defaults with user overrides
    auto_filled: dict[str, Any] = {f: FIELD_DEFAULTS[f] for f in missing_fields if f in FIELD_DEFAULTS}
    if default_overrides:
        for f, val in default_overrides.items():
            if val is not None and str(val).strip():
                auto_filled[f] = val

    created, updated, skipped, errors = 0, 0, 0, []
    error_records: list[dict[str, Any]] = []

    for rel_idx, row in enumerate(rows_2d[data_start_index:]):
        excel_row = data_start_index + rel_idx + 1
        try:
            name = _cell(row, col_name)
            if not name and (col_first is not None or col_last is not None):
                p1 = _cell(row, col_first) if col_first is not None else ""
                p2 = _cell(row, col_last) if col_last is not None else ""
                name = " ".join(p for p in (p1, p2) if p).strip()
            name = normalize_product_name(name)
            if not name:
                continue

            price_s = _cell(row, col_price) if col_price is not None else str(auto_filled.get("price", "0"))
            qty_s = _cell(row, col_qty) if col_qty is not None else str(auto_filled.get("quantity", "0"))
            thr_s = _cell(row, col_thr) if col_thr is not None else str(auto_filled.get("stock_threshold", "5"))
            price = parse_price(price_s)
            quantity = parse_price(qty_s)  # int cỡ số nguyên
            threshold = parse_price(thr_s)
            if quantity < 0:
                quantity = 0
            if threshold < 0:
                threshold = 5

            desc = _cell(row, col_desc) if col_desc is not None else str(auto_filled.get("description", ""))
            cell_img_url = (
                _cell(row, col_image) if col_image is not None else str(auto_filled.get("image_url", ""))
            )
            row_0based_excel = data_start_index + rel_idx
            extracted_url = ""
            if extracted_images:
                extracted_url = extracted_images.get(row_0based_excel, "") or ""
            img_url = extracted_url or cell_img_url
            cat = _cell(row, col_cat) if col_cat is not None else str(auto_filled.get("category", ""))
            # Phase 3: AI categories (suggestions or manual overrides) take priority
            if ai_categories and name in ai_categories:
                cat = ai_categories[name]
            
            # --- Phase 4: Manual Overrides (Highest Priority) ---
            if manual_overrides and name in manual_overrides:
                ovr = manual_overrides[name]
                if "category" in ovr: cat = ovr["category"]
                if "price" in ovr: price = parse_price(ovr["price"])
                if "quantity" in ovr: quantity = parse_price(ovr["quantity"])
                if "stock_threshold" in ovr: threshold = parse_price(ovr["stock_threshold"])
                if "description" in ovr: desc = ovr["description"]
                if "image_url" in ovr: img_url = ovr["image_url"]
            extra = _row_import_extra(row, headers_row, column_mapping)
            meta_payload = None
            if extra:
                meta_payload = {"import_extra": extra}

            existing = None
            if duplicate_strategy in ("skip", "update"):
                stmt = select(Product).where(
                    Product.workspace_id == workspace_id,
                    Product.name == name,
                )
                result = await db.execute(stmt)
                existing = result.scalar_one_or_none()

            if existing and duplicate_strategy == "skip":
                skipped += 1
                continue

            if existing and duplicate_strategy == "update":
                existing.price = price
                existing.quantity = quantity
                existing.stock_threshold = threshold
                if desc:
                    existing.description = desc
                if img_url:
                    existing.image_url = img_url
                if cat:
                    existing.category = cat
                if meta_payload is not None:
                    md = dict(existing.metadata_json or {})
                    prev = dict(md.get("import_extra") or {})
                    prev.update(meta_payload["import_extra"])
                    md["import_extra"] = prev
                    existing.metadata_json = md
                updated += 1
                if not dry_run and existing.quantity <= existing.stock_threshold:
                    await NotificationService.notify_low_stock(
                        workspace_id,
                        existing.name,
                        existing.quantity,
                        existing.stock_threshold,
                    )
            else:
                product = Product(
                    workspace_id=workspace_id,
                    name=name,
                    description=desc or "",
                    category=cat or None,
                    price=price,
                    quantity=quantity,
                    stock_threshold=threshold,
                    image_url=img_url or None,
                    metadata_json=meta_payload,
                )
                db.add(product)
                created += 1
        except Exception as e:
            msg = f"Dòng {excel_row}: {e}"
            errors.append(msg)
            error_records.append({"row": excel_row, "message": str(e)})
            if len(errors) <= error_sample_limit:
                pass

    await _after_import_commit(db, dry_run, workspace_id)

    data_rows = max(0, len(rows_2d) - data_start_index)
    err_sample = errors[:error_sample_limit]
    err_csv = build_import_error_csv(error_records) if error_records else ""
    return {
        "total_rows": data_rows,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": err_sample,
        "errors_total": len(errors),
        "error_csv": err_csv,
        "dry_run": dry_run,
        # Phase 1: smart-defaults metadata
        "missing_fields": missing_fields,
        "auto_filled_fields": auto_filled,
    }


async def import_customers_from_values(
    db: AsyncSession,
    workspace_id: uuid.UUID,
    rows_2d: list[list[Any]],
    header_row_index: int = 0,
    data_start_index: int = 1,
    column_mapping: dict[str, str] | None = None,
    *,
    duplicate_strategy: DuplicateStrategy = "update",
    dry_run: bool = False,
    error_sample_limit: int = 50,
) -> dict[str, Any]:
    if header_row_index >= len(rows_2d):
        raise ValueError("Không có dòng tiêu đề.")

    headers_raw = rows_2d[header_row_index]
    headers_str = [str(h or "") for h in headers_raw]
    hidx = _build_header_index(headers_str)

    col_name = _find_col(
        hidx,
        NAME_KEYS + ("họ tên", "ho ten", "fullname", "full name"),
        "name",
        column_mapping,
    )
    col_phone = _find_col(hidx, PHONE_KEYS, "phone", column_mapping)
    col_email = _find_col(hidx, EMAIL_KEYS, "email", column_mapping)
    col_note = _find_col(hidx, NOTE_KEYS, "note", column_mapping)
    col_first = _find_col(hidx, FIRST_NAME_KEYS, "first_name", column_mapping)
    col_last = _find_col(hidx, LAST_NAME_KEYS, "last_name", column_mapping)

    if col_phone is None and col_email is None:
        raise ValueError("Cần ít nhất một cột số điện thoại hoặc email để khớp khách hàng.")

    created, updated, skipped, errors = 0, 0, 0, []
    error_records: list[dict[str, Any]] = []

    for rel_idx, row in enumerate(rows_2d[data_start_index:]):
        excel_row = data_start_index + rel_idx + 1
        try:
            phone = normalize_phone(_cell(row, col_phone)) if col_phone is not None else None
            email = normalize_email(_cell(row, col_email)) if col_email is not None else None
            name = normalize_product_name(_concat_customer_name(row, col_name, col_first, col_last, column_mapping))
            note = _cell(row, col_note) if col_note is not None else ""
            extra = _row_import_extra(row, headers_str, column_mapping)
            meta_note: dict[str, Any] = {}
            if note:
                meta_note["import_note"] = note
            if extra:
                meta_note["import_extra"] = extra

            if not phone and not email:
                msg = f"Dòng {excel_row}: thiếu SĐT và email hợp lệ"
                errors.append(msg)
                error_records.append({"row": excel_row, "message": msg.replace(f"Dòng {excel_row}: ", "")})
                continue

            existing = None
            if duplicate_strategy in ("skip", "update"):
                if phone:
                    r = await db.execute(
                        select(Customer).where(
                            Customer.workspace_id == workspace_id,
                            Customer.phone == phone,
                        )
                    )
                    existing = r.scalar_one_or_none()
                if not existing and email:
                    r = await db.execute(
                        select(Customer).where(
                            Customer.workspace_id == workspace_id,
                            Customer.email == email,
                        )
                    )
                    existing = r.scalar_one_or_none()

            if existing and duplicate_strategy == "skip":
                skipped += 1
                continue

            if existing and duplicate_strategy == "update":
                if name:
                    existing.name = name
                if email and not existing.email:
                    existing.email = email
                if phone and not existing.phone:
                    existing.phone = phone
                if meta_note:
                    md = dict(existing.metadata_json or {})
                    if "import_note" in meta_note:
                        md["import_note"] = meta_note["import_note"]
                    if "import_extra" in meta_note:
                        prev = dict(md.get("import_extra") or {})
                        prev.update(meta_note["import_extra"])
                        md["import_extra"] = prev
                    existing.metadata_json = md
                updated += 1
            else:
                cust = Customer(
                    workspace_id=workspace_id,
                    phone=phone,
                    email=email,
                    name=name or None,
                    metadata_json=meta_note if meta_note else None,
                )
                db.add(cust)
                created += 1
        except Exception as e:
            msg = f"Dòng {excel_row}: {e}"
            errors.append(msg)
            error_records.append({"row": excel_row, "message": str(e)})

    await _after_import_commit(db, dry_run, workspace_id)

    data_rows = max(0, len(rows_2d) - data_start_index)
    err_sample = errors[:error_sample_limit]
    err_csv = build_import_error_csv(error_records) if error_records else ""
    return {
        "total_rows": data_rows,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": err_sample,
        "errors_total": len(errors),
        "error_csv": err_csv,
        "dry_run": dry_run,
    }


async def run_import_for_workspace(
    db: AsyncSession,
    workspace: Workspace,
    spreadsheet_id: str,
    sheet_name: str,
    entity: str,
    header_row_1based: int = 1,
    data_start_row_1based: int = 2,
    range_a1: str | None = None,
    column_mapping: dict[str, str] | None = None,
    *,
    duplicate_strategy: DuplicateStrategy = "update",
    dry_run: bool = False,
    ai_categories: dict[str, str] | None = None,
    default_overrides: dict[str, Any] | None = None,
    manual_overrides: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    spreadsheet_id = _extract_spreadsheet_id(spreadsheet_id)
    values = await fetch_sheet_values(db, workspace, spreadsheet_id, sheet_name, range_a1)
    h_idx = max(0, header_row_1based - 1)
    d_idx = max(h_idx + 1, data_start_row_1based - 1)

    entity_norm = (entity or "products").lower()

    extracted_images: dict[int, str] | None = None
    products_entity = entity_norm not in ("customer", "customers", "khach", "khách")
    if products_entity and not dry_run:
        try:
            xlsx_blob = await fetch_spreadsheet_export_xlsx_bytes(
                db,
                workspace,
                spreadsheet_id,
            )
            from app.services.image_extractor import extract_images_from_xlsx, upload_and_map

            raw_images = extract_images_from_xlsx(
                xlsx_blob,
                data_start_row_0based=d_idx,
                worksheet_title=sheet_name,
            )
            if raw_images:
                extracted_images = await upload_and_map(raw_images, str(workspace.id))
                logger.info(
                    "Google Sheets (Drive export): extracted %d images, uploaded %d (tab=%r)",
                    len(raw_images),
                    len(extracted_images or {}),
                    sheet_name,
                )
        except ValueError:
            # Re-raise known errors (like 403 Forbidden for missing Drive scopes)
            # so the background job fails and the user sees the message in the UI.
            # UPDATE: As requested, we swallow this error to allow text-only import to continue
            # without requiring the user to have drive.readonly scopes approved immediately.
            logger.warning(
                "Google Sheet embedded-image sync skipped (Drive export/XLSX) due to ValueError (likely 403 Forbidden); "
                "text import continues. Reconnect Google if Drive scope missing.",
                exc_info=True,
            )
        except Exception:
            logger.warning(
                "Google Sheet embedded-image sync skipped (Drive export/XLSX); "
                "text import continues. Reconnect Google if Drive scope missing.",
                exc_info=True,
            )

    if entity_norm in ("customer", "customers", "khach", "khách"):
        return await import_customers_from_values(
            db,
            workspace.id,
            values,
            h_idx,
            d_idx,
            column_mapping,
            duplicate_strategy=duplicate_strategy,
            dry_run=dry_run,
        )
    return await import_products_from_values(
        db,
        workspace.id,
        values,
        h_idx,
        d_idx,
        column_mapping,
        duplicate_strategy=duplicate_strategy,
        dry_run=dry_run,
        extracted_images=extracted_images,
        ai_categories=ai_categories,
        default_overrides=default_overrides,
        manual_overrides=manual_overrides,
    )


async def import_from_file_bytes(
    db: AsyncSession,
    workspace_id: uuid.UUID,
    content: bytes,
    filename: str,
    entity: str,
    header_row_1based: int = 1,
    data_start_row_1based: int = 2,
    column_mapping: dict[str, str] | None = None,
    *,
    duplicate_strategy: DuplicateStrategy = "update",
    dry_run: bool = False,
    ai_categories: dict[str, str] | None = None,
    default_overrides: dict[str, Any] | None = None,
    manual_overrides: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    values = parse_upload_to_rows(content, filename, max_rows=None)
    h_idx = max(0, header_row_1based - 1)
    d_idx = max(h_idx + 1, data_start_row_1based - 1)
    entity_norm = (entity or "products").lower()
    if entity_norm in ("customer", "customers", "khach", "khách"):
        return await import_customers_from_values(
            db,
            workspace_id,
            values,
            h_idx,
            d_idx,
            column_mapping,
            duplicate_strategy=duplicate_strategy,
            dry_run=dry_run,
        )

    # --- Phase 2: extract & upload embedded images (XLSX only) ---
    extracted_images: dict[int, str] | None = None
    name_lower = (filename or "").lower()
    if name_lower.endswith(".xlsx") and not dry_run:
        try:
            from app.services.image_extractor import (
                extract_images_from_xlsx,
                upload_and_map,
            )
            raw_images = extract_images_from_xlsx(
                content, data_start_row_0based=d_idx,
            )
            if raw_images:
                extracted_images = await upload_and_map(
                    raw_images, str(workspace_id),
                )
                logger.info(
                    "Extracted %d images, uploaded %d for workspace %s",
                    len(raw_images),
                    len(extracted_images or {}),
                    workspace_id,
                )
        except Exception:
            logger.warning(
                "Image extraction failed for %s — continuing without images",
                filename,
                exc_info=True,
            )

    return await import_products_from_values(
        db,
        workspace_id,
        values,
        h_idx,
        d_idx,
        column_mapping,
        duplicate_strategy=duplicate_strategy,
        dry_run=dry_run,
        extracted_images=extracted_images,
        ai_categories=ai_categories,
        default_overrides=default_overrides,
        manual_overrides=manual_overrides,
    )

